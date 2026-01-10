"""
Flask backend for Biometric Attendance System (Fingerprint-Only).
Handles teacher registration and attendance recording via fingerprint authentication.
Supports mode-based operation: REGISTER MODE and ATTENDANCE MODE.
"""
from flask import Flask, request, render_template, jsonify, make_response, send_file
from datetime import datetime, timedelta
import pytz
import database
from config import COOLDOWN_MINUTES
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import os

# Bangladesh timezone (UTC+6)
BD_TIMEZONE = pytz.timezone('Asia/Dhaka')

app = Flask(__name__)

# In-memory storage for latest scanned fingerprint_id (waiting for registration)
# Structure: { "fingerprint_id": int, "timestamp": ISO8601 string }
_latest_fingerprint_id = None


def calculate_working_hours(check_in_time, check_out_time):
    """
    Calculate working hours between check-in and check-out times.
    
    Args:
        check_in_time: Check-in time string (HH:MM:SS)
        check_out_time: Check-out time string (HH:MM:SS)
    
    Returns:
        str: Human-readable working hours (e.g., "8 hours 30 minutes")
    """
    try:
        # Parse time strings
        check_in = datetime.strptime(check_in_time, "%H:%M:%S").time()
        check_out = datetime.strptime(check_out_time, "%H:%M:%S").time()
        
        # Create datetime objects for calculation (using BD timezone)
        today = datetime.now(BD_TIMEZONE).date()
        check_in_dt = datetime.combine(today, check_in)
        check_out_dt = datetime.combine(today, check_out)
        
        # Handle case where check-out is next day
        if check_out_dt < check_in_dt:
            check_out_dt += timedelta(days=1)
        
        # Calculate difference
        diff = check_out_dt - check_in_dt
        total_seconds = int(diff.total_seconds())
        
        hours = total_seconds // 3600
        minutes = (total_seconds % 3600) // 60
        
        return f"{hours} hours {minutes} minutes"
    except Exception as e:
        print(f"Error calculating working hours: {e}")
        return "0 hours 0 minutes"


def parse_working_hours_to_minutes(working_hours_str):
    """
    Parse working hours string (e.g., "8 hours 30 minutes") to total minutes.
    
    Args:
        working_hours_str: Working hours string like "8 hours 30 minutes"
    
    Returns:
        int: Total minutes, or 0 if parsing fails
    """
    try:
        if not working_hours_str:
            return 0
        
        # Extract hours and minutes from string like "8 hours 30 minutes"
        import re
        hours_match = re.search(r'(\d+)\s*hours?', working_hours_str)
        minutes_match = re.search(r'(\d+)\s*minutes?', working_hours_str)
        
        hours = int(hours_match.group(1)) if hours_match else 0
        minutes = int(minutes_match.group(1)) if minutes_match else 0
        
        return hours * 60 + minutes
    except Exception as e:
        print(f"Error parsing working hours: {e}")
        return 0


def format_minutes_to_hours(minutes):
    """
    Convert total minutes to human-readable format.
    
    Args:
        minutes: Total minutes
    
    Returns:
        str: Formatted string like "8 hours 30 minutes"
    """
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours} hours {mins} minutes"


def get_server_time():
    """Get current server time in ISO8601 format (Bangladesh timezone)."""
    return datetime.now(BD_TIMEZONE).isoformat()


def get_date_string():
    """Get current date in YYYY-MM-DD format (Bangladesh timezone)."""
    return datetime.now(BD_TIMEZONE).strftime("%Y-%m-%d")


def get_time_string():
    """Get current time in HH:MM:SS format (Bangladesh timezone)."""
    return datetime.now(BD_TIMEZONE).strftime("%H:%M:%S")


@app.route('/')
def index():
    """Admin page with mode control and registration form."""
    current_mode = database.get_system_mode()
    return render_template('register.html', current_mode=current_mode)


@app.route('/mode', methods=['GET'])
def get_mode():
    """
    Get current system mode.
    ESP32 can query this to know which mode is active.
    """
    try:
        current_mode = database.get_system_mode()
        return jsonify({
            'status': 'success',
            'mode': current_mode,
            'server_time': get_server_time()
        }), 200
    except Exception as e:
        print(f"Error getting mode: {e}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to get system mode: {str(e)}',
            'server_time': get_server_time()
        }), 500


@app.route('/mode', methods=['POST'])
def set_mode():
    """
    Set system mode (admin only).
    Request: { "mode": "register" | "attendance" }
    """
    try:
        if not request.is_json:
            return jsonify({
                'status': 'error',
                'message': 'Request must be JSON',
                'server_time': get_server_time()
            }), 400
        
        data = request.get_json()
        mode = data.get('mode', '').lower()
        
        if mode not in ['register', 'attendance']:
            return jsonify({
                'status': 'error',
                'message': 'Mode must be "register" or "attendance"',
                'server_time': get_server_time()
            }), 400
        
        success = database.set_system_mode(mode)
        
        if success:
            return jsonify({
                'status': 'success',
                'message': f'System mode set to {mode}',
                'mode': mode,
                'server_time': get_server_time()
            }), 200
        else:
            return jsonify({
                'status': 'error',
                'message': 'Failed to set system mode',
                'server_time': get_server_time()
            }), 500
    
    except Exception as e:
        print(f"Error setting mode: {e}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to set mode: {str(e)}',
            'server_time': get_server_time()
        }), 500


@app.route('/register', methods=['GET', 'POST'])
def register():
    """
    Register a new teacher - saves all three fields together.
    
    Request (JSON or form-data):
    {
        "name": "John Doe",
        "department": "CSE",
        "fingerprint_id": 7
    }
    
    In REGISTER MODE:
    - Accepts name, department, and fingerprint_id together
    - Validates all fields are present
    - Saves complete teacher record
    
    In ATTENDANCE MODE:
    - Returns error (registration disabled)
    """
    if request.method == 'GET':
        current_mode = database.get_system_mode()
        return render_template('register.html', current_mode=current_mode)
    
    try:
        # Check system mode
        current_mode = database.get_system_mode()
        
        if current_mode != 'register':
            return jsonify({
                'status': 'error',
                'message': f'System is in {current_mode} mode. Switch to register mode to register teachers.'
            }), 403
        
        # Get data (supports both JSON and form-data)
        if request.is_json:
            data = request.get_json()
            name = data.get('name', '').strip()
            department = data.get('department', '').strip()
            fingerprint_id = data.get('fingerprint_id')
        else:
            name = request.form.get('name', '').strip()
            department = request.form.get('department', '').strip()
            fingerprint_id_str = request.form.get('fingerprint_id', '').strip()
            fingerprint_id = int(fingerprint_id_str) if fingerprint_id_str else None
        
        # Validate all required fields
        if not name or not department or fingerprint_id is None:
            return jsonify({
                'status': 'error',
                'message': 'Missing required fields: name, department, fingerprint_id'
            }), 400
        
        # Validate fingerprint_id type
        try:
            fingerprint_id = int(fingerprint_id)
        except (ValueError, TypeError):
            return jsonify({
                'status': 'error',
                'message': 'Invalid fingerprint_id format (must be integer)'
            }), 400
        
        # Check if fingerprint_id already exists
        existing_teacher = database.get_teacher_by_fingerprint_id(fingerprint_id)
        if existing_teacher:
            return jsonify({
                'status': 'error',
                'message': f'Fingerprint ID {fingerprint_id} already registered'
            }), 400
        
        # Generate unique teacher_id
        teacher_id = f"teacher_{datetime.now(BD_TIMEZONE).strftime('%Y%m%d%H%M%S')}_{fingerprint_id}"
        
        # Register teacher in database (all three fields together)
        success = database.register_teacher(
            teacher_id=teacher_id,
            name=name,
            department=department,
            fingerprint_id=fingerprint_id
        )
        
        if success:
            # Clear the latest fingerprint_id after successful registration
            global _latest_fingerprint_id
            _latest_fingerprint_id = None
            
            return jsonify({
                'status': 'success',
                'message': f'Teacher {name} registered successfully',
                'teacher_id': teacher_id
            }), 201
        else:
            return jsonify({
                'status': 'error',
                'message': 'Failed to register teacher in database'
            }), 500
    
    except Exception as e:
        print(f"Registration error: {e}")
        return jsonify({
            'status': 'error',
            'message': f'Registration failed: {str(e)}'
        }), 500


@app.route('/register-fingerprint', methods=['POST'])
def register_fingerprint():
    """
    ESP32 sends scanned fingerprint_id (REGISTER MODE only).
    
    ESP32 sends:
    {
        "fingerprint_id": 7
    }
    
    Backend:
    - Validates system is in REGISTER MODE
    - Stores fingerprint_id in memory (latest_fingerprint_id)
    - Rejects duplicate fingerprint_id (already registered)
    - Returns success with fingerprint_id
    """
    try:
        server_time_iso = get_server_time()
        
        # Check system mode
        current_mode = database.get_system_mode()
        
        if current_mode != 'register':
            return jsonify({
                'status': 'error',
                'message': f'System is in {current_mode} mode',
                'server_time': server_time_iso
            }), 403
        
        # Validate request format
        if not request.is_json:
            return jsonify({
                'status': 'error',
                'message': 'Request must be JSON',
                'server_time': server_time_iso
            }), 400
        
        data = request.get_json()
        fingerprint_id = data.get('fingerprint_id')
        
        # Validate fingerprint_id
        if fingerprint_id is None:
            return jsonify({
                'status': 'error',
                'message': 'Missing fingerprint_id in request',
                'server_time': server_time_iso
            }), 400
        
        try:
            fingerprint_id = int(fingerprint_id)
        except (ValueError, TypeError):
            return jsonify({
                'status': 'error',
                'message': 'Invalid fingerprint_id format (must be integer)',
                'server_time': server_time_iso
            }), 400
        
        # Check if fingerprint_id already exists (already registered)
        existing_teacher = database.get_teacher_by_fingerprint_id(fingerprint_id)
        if existing_teacher:
            return jsonify({
                'status': 'error',
                'message': f'Fingerprint ID {fingerprint_id} already registered',
                'server_time': server_time_iso
            }), 400
        
        # Store fingerprint_id in memory (waiting for registration form submission)
        global _latest_fingerprint_id
        _latest_fingerprint_id = {
            'fingerprint_id': fingerprint_id,
            'timestamp': get_server_time()
        }
        
        return jsonify({
            'status': 'success',
            'message': 'Fingerprint ID received and stored',
            'fingerprint_id': fingerprint_id,
            'server_time': server_time_iso
        }), 200
    
    except Exception as e:
        print(f"Fingerprint registration error: {e}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to store fingerprint: {str(e)}',
            'server_time': get_server_time()
        }), 500


@app.route('/register-fingerprint/latest', methods=['GET'])
def get_latest_fingerprint():
    """
    Frontend polls this endpoint to check if fingerprint_id is available.
    
    Returns:
    {
        "status": "waiting" | "ready",
        "fingerprint_id": 7  // Only present when status == "ready"
    }
    """
    try:
        global _latest_fingerprint_id
        
        if _latest_fingerprint_id is None:
            return jsonify({
                'status': 'waiting',
                'message': 'No fingerprint scanned yet'
            }), 200
        else:
            return jsonify({
                'status': 'ready',
                'fingerprint_id': _latest_fingerprint_id['fingerprint_id'],
                'timestamp': _latest_fingerprint_id['timestamp']
            }), 200
    
    except Exception as e:
        print(f"Error getting latest fingerprint: {e}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to get fingerprint status: {str(e)}'
        }), 500


@app.route('/register-fingerprint/clear', methods=['POST'])
def clear_latest_fingerprint():
    """
    Clear the stored fingerprint_id (optional endpoint).
    Can be called after successful registration or to reset.
    """
    try:
        global _latest_fingerprint_id
        _latest_fingerprint_id = None
        
        return jsonify({
            'status': 'success',
            'message': 'Fingerprint ID cleared'
        }), 200
    
    except Exception as e:
        print(f"Error clearing fingerprint: {e}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to clear fingerprint: {str(e)}'
        }), 500


@app.route('/attendance', methods=['POST'])
def attendance():
    """
    Main attendance endpoint - Fingerprint-only authentication.
    
    ESP32 sends: { "fingerprint_id": int }
    Backend automatically decides: check-in or check-out
    
    Attendance Logic (UNCHANGED):
    1. First scan of day → check_in
    2. Second scan (≥15 min after check-in) → check_out
    3. Early checkout (<15 min) → reject with remaining_minutes
    4. Duplicate scan (already checked out) → reject
    
    All timestamps use server time (ESP32 time ignored).
    
    Mode Enforcement:
    - Only works in ATTENDANCE MODE
    - Returns error if in REGISTER MODE
    """
    try:
        # Get current server time and date (authoritative)
        current_time = get_time_string()
        current_date = get_date_string()
        server_time_iso = get_server_time()
        
        # Check system mode
        current_mode = database.get_system_mode()
        
        if current_mode != 'attendance':
            return jsonify({
                'status': 'error',
                'message': f'System is in {current_mode} mode',
                'server_time': server_time_iso
            }), 403
        
        # Validate request format
        if not request.is_json:
            return jsonify({
                'status': 'error',
                'message': 'Request must be JSON with fingerprint_id',
                'server_time': server_time_iso
            }), 400
        
        data = request.get_json()
        fingerprint_id = data.get('fingerprint_id')
        
        if fingerprint_id is None:
            return jsonify({
                'status': 'error',
                'message': 'Missing fingerprint_id in request',
                'server_time': server_time_iso
            }), 400
        
        # Validate fingerprint_id type
        try:
            fingerprint_id = int(fingerprint_id)
        except (ValueError, TypeError):
            return jsonify({
                'status': 'error',
                'message': 'Invalid fingerprint_id format (must be integer)',
                'server_time': server_time_iso
            }), 400
        
        # Find teacher by fingerprint_id
        teacher = database.get_teacher_by_fingerprint_id(fingerprint_id)
        
        if not teacher:
            return jsonify({
                'status': 'error',
                'action': 'not_found',
                'message': f'Fingerprint ID {fingerprint_id} not registered',
                'teacher': None,
                'server_time': server_time_iso
            }), 404
        
        # Teacher identified - proceed with attendance logic
        teacher_id = teacher['teacher_id']
        teacher_name = teacher.get('name')
        department = teacher.get('department')
        
        # Get today's attendance record
        today_attendance = database.get_today_attendance(teacher_id, current_date)
        
        # ============================================
        # ATTENDANCE DECISION LOGIC (UNCHANGED)
        # ============================================
        
        # Case 1: No record exists → create check_in
        if not today_attendance:
            success = database.create_check_in(teacher_id, current_date, current_time)
            if success:
                return jsonify({
                    'status': 'success',
                    'action': 'check_in',
                    'message': 'Checked in successfully',
                    'teacher': {
                        'name': teacher_name,
                        'department': department
                    },
                    'check_in': current_time,
                    'oled': [
                        f"Welcome {teacher_name}",
                        f"Dept: {department}",
                        "Checked in",
                        current_time
                    ],
                    'server_time': server_time_iso
                }), 200
            else:
                return jsonify({
                    'status': 'error',
                    'action': 'error',
                    'message': 'Failed to record check-in',
                    'teacher': {
                        'name': teacher_name,
                        'department': department
                    },
                    'server_time': server_time_iso
                }), 500
        
        # Case 2: check_in exists, check_out does NOT exist
        check_in = today_attendance.get('check_in')
        check_out = today_attendance.get('check_out')
        
        if check_in and not check_out:
            # Attempt check-out, but enforce 15-minute cooldown
            
            try:
                # Parse check-in time
                check_in_time = datetime.strptime(check_in, "%H:%M:%S").time()
                current_time_obj = datetime.strptime(current_time, "%H:%M:%S").time()
                
                # Create datetime objects for comparison (using BD timezone)
                today = datetime.now(BD_TIMEZONE).date()
                check_in_dt = datetime.combine(today, check_in_time)
                current_dt = datetime.combine(today, current_time_obj)
                
                # Handle case where current time is next day (midnight crossover)
                if current_dt < check_in_dt:
                    current_dt += timedelta(days=1)
                
                # Calculate time difference in minutes
                time_diff = current_dt - check_in_dt
                minutes_passed = int(time_diff.total_seconds() / 60)
                
                # Enforce cooldown: reject if < 15 minutes
                if minutes_passed < COOLDOWN_MINUTES:
                    remaining_minutes = COOLDOWN_MINUTES - minutes_passed
                    return jsonify({
                        'status': 'error',
                        'action': 'cooldown',
                        'message': f'Please try again after {remaining_minutes} minute(s)',
                        'remaining_minutes': remaining_minutes,
                        'teacher': {
                            'name': teacher_name,
                            'department': department,
                            'check_in': check_in
                        },
                        'oled': [
                            f"{teacher_name}",
                            "Please wait",
                            f"{remaining_minutes} min left"
                        ],
                        'server_time': server_time_iso
                    }), 400
                
                # Valid check-out (≥15 minutes passed)
                working_hours = calculate_working_hours(check_in, current_time)
                success = database.create_check_out(
                    teacher_id, current_date, current_time, working_hours
                )
                
                if success:
                    return jsonify({
                        'status': 'success',
                        'action': 'check_out',
                        'message': 'Checked out successfully',
                        'teacher': {
                            'name': teacher_name,
                            'department': department
                        },
                        'check_in': check_in,
                        'check_out': current_time,
                        'working_hours': working_hours,
                        'oled': [
                            f"Goodbye {teacher_name}",
                            "Checked out",
                            working_hours
                        ],
                        'server_time': server_time_iso
                    }), 200
                else:
                    return jsonify({
                        'status': 'error',
                        'action': 'error',
                        'message': 'Failed to record check-out',
                        'teacher': {
                            'name': teacher_name,
                            'department': department
                        },
                        'server_time': server_time_iso
                    }), 500
            
            except Exception as e:
                print(f"Error processing check-out: {e}")
                return jsonify({
                    'status': 'error',
                    'message': 'Error processing check-out time',
                    'server_time': server_time_iso
                }), 500
        
        # Case 3: Both check_in and check_out exist → reject (duplicate scan)
        elif check_in and check_out:
            return jsonify({
                'status': 'error',
                'action': 'completed',
                'message': 'Attendance already completed for today',
                'teacher': {
                    'name': teacher_name,
                    'department': department
                },
                'oled': [
                    "Attendance done",
                    "Come tomorrow"
                ],
                'server_time': server_time_iso
            }), 400
        
        # Case 4: Unexpected state (should not happen)
        else:
            return jsonify({
                'status': 'error',
                'action': 'error',
                'message': 'Invalid attendance state',
                'teacher': {
                    'name': teacher_name,
                    'department': department
                },
                'server_time': server_time_iso
            }), 500
    
    except Exception as e:
        print(f"Attendance error: {e}")
        return jsonify({
            'status': 'error',
            'action': 'error',
            'message': f'Attendance processing failed: {str(e)}',
            'server_time': get_server_time()
        }), 500


@app.route('/teachers', methods=['GET'])
def get_teachers():
    """
    Get all teachers and their attendance records.
    Used by frontend to display attendance table.
    
    Query parameters:
    - start_date: Optional filter start date (YYYY-MM-DD)
    - end_date: Optional filter end date (YYYY-MM-DD)
    """
    try:
        # Get date range filters
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        
        teachers = database.get_all_teachers()
        # Flatten attendance into records for easier UI rendering
        records = []
        for teacher_id, data in teachers.items():
            name = data.get('name')
            department = data.get('department')
            attendance = data.get('attendance', {}) or {}
            if not attendance:
                # Only include if no date filter is applied
                if not start_date and not end_date:
                    records.append({
                        'teacher_id': teacher_id,
                        'name': name,
                        'department': department,
                        'date': None,
                        'check_in': None,
                        'check_out': None,
                        'working_hours': None
                    })
            else:
                for date_str, rec in attendance.items():
                    # Apply date range filter
                    if start_date or end_date:
                        if start_date and date_str < start_date:
                            continue
                        if end_date and date_str > end_date:
                            continue
                    
                    records.append({
                        'teacher_id': teacher_id,
                        'name': name,
                        'department': department,
                        'date': date_str,
                        'check_in': rec.get('check_in'),
                        'check_out': rec.get('check_out'),
                        'working_hours': rec.get('working_hours')
                    })

        return jsonify({
            'status': 'success',
            'records': records,
            'filters': {
                'start_date': start_date,
                'end_date': end_date
            },
            'server_time': get_server_time()
        }), 200
    except Exception as e:
        print(f"Error getting teachers: {e}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to get teachers: {str(e)}',
            'server_time': get_server_time()
        }), 500


@app.route('/attendance/download', methods=['GET'])
def download_attendance_excel():
    """
    Download attendance records as Excel file with multi-sheet format.
    Sheet 1: Summary Report (overall statistics by teacher)
    Sheet 2+: Daily sheets (one per day in date range)
    
    Query parameters:
    - start_date: Optional filter start date (YYYY-MM-DD)
    - end_date: Optional filter end date (YYYY-MM-DD)
    """
    try:
        # Get date range filters (strip whitespace and handle empty strings)
        start_date = request.args.get('start_date', '').strip() or None
        end_date = request.args.get('end_date', '').strip() or None
        
        teachers = database.get_all_teachers()
        
        # Debug: Print how many teachers we got
        print(f"Excel Download: Found {len(teachers)} teachers")
        
        # Create a new workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Collect all records and organize by date
        all_records = []
        dates_set = set()
        
        for teacher_id, data in teachers.items():
            name = data.get('name', '')
            department = data.get('department', '')
            attendance = data.get('attendance', {}) or {}
            
            # Debug: Print attendance data for this teacher
            if attendance:
                print(f"Teacher {name}: Found {len(attendance)} attendance records")
            
            for date_str, rec in attendance.items():
                # Apply date range filter (only if dates are provided)
                if start_date:
                    if date_str < start_date:
                        continue
                if end_date:
                    if date_str > end_date:
                        continue
                
                check_in = rec.get('check_in')
                check_out = rec.get('check_out')
                working_hours = rec.get('working_hours')
                
                # Include records with both check-in and check-out (for summary)
                # For daily sheets, we'll include all records
                if check_in and check_out:
                    all_records.append({
                        'teacher_id': teacher_id,
                        'name': name,
                        'department': department,
                        'date': date_str,
                        'check_in': check_in,
                        'check_out': check_out,
                        'working_hours': working_hours
                    })
                    dates_set.add(date_str)
        
        # Also collect all records (including partial) for daily sheets
        all_records_for_daily = []
        for teacher_id, data in teachers.items():
            name = data.get('name', '')
            department = data.get('department', '')
            attendance = data.get('attendance', {}) or {}
            
            for date_str, rec in attendance.items():
                # Apply date range filter
                if start_date:
                    if date_str < start_date:
                        continue
                if end_date:
                    if date_str > end_date:
                        continue
                
                check_in = rec.get('check_in')
                check_out = rec.get('check_out')
                working_hours = rec.get('working_hours')
                
                # Include all records for daily sheets (even if only check-in)
                if check_in:  # At least check-in is required
                    all_records_for_daily.append({
                        'teacher_id': teacher_id,
                        'name': name,
                        'department': department,
                        'date': date_str,
                        'check_in': check_in,
                        'check_out': check_out,
                        'working_hours': working_hours
                    })
                    dates_set.add(date_str)
        
        # Debug: Print how many records we collected
        print(f"Excel Download: Collected {len(all_records)} complete records")
        print(f"Excel Download: Collected {len(all_records_for_daily)} total records for daily sheets")
        print(f"Excel Download: Found {len(dates_set)} unique dates")
        print(f"Excel Download: Date range filter: {start_date} to {end_date}")
        
        # Sort dates
        sorted_dates = sorted(dates_set) if dates_set else []
        
        # Define header styling (used for both summary and daily sheets)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # ============================================
        # SHEET 1: SUMMARY REPORT
        # ============================================
        summary_ws = wb.create_sheet("Summary Report", 0)
        
        # If no records at all, add a message
        if not all_records_for_daily:
            summary_ws.append(['No attendance records found for the selected date range.'])
            summary_ws.column_dimensions['A'].width = 50
        else:
            # Summary headers (removed date columns as requested)
            summary_headers = [
                'Teacher Name', 'Department', 'Total Days Worked', 
                'Total Working Hours', 'Average Hours per Day'
            ]
            summary_ws.append(summary_headers)
            
            # Style summary header
            for cell in summary_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Calculate summary statistics for each teacher
            # Use all_records (complete records only) for accurate statistics
            teacher_summary = {}
            
            # Process complete records only (from all_records, not all_records_for_daily)
            for rec in all_records:
                teacher_id = rec['teacher_id']
                check_in = rec.get('check_in', '').strip() if rec.get('check_in') else ''
                check_out = rec.get('check_out', '').strip() if rec.get('check_out') else ''
                working_hours = rec.get('working_hours', '').strip() if rec.get('working_hours') else ''
                
                # Only process records with both check-in and check-out
                if check_in and check_out:
                    if teacher_id not in teacher_summary:
                        teacher_summary[teacher_id] = {
                            'name': rec['name'],
                            'department': rec['department'],
                            'days': 0,
                            'total_minutes': 0
                        }
                    
                    summary = teacher_summary[teacher_id]
                    summary['days'] += 1
                    
                    # Parse working hours and add to total
                    minutes = parse_working_hours_to_minutes(working_hours)
                    summary['total_minutes'] += minutes
                    
                    # Debug output
                    print(f"Summary: {rec['name']} - Date: {rec['date']}, Working Hours: {working_hours}, Minutes: {minutes}")
            
            # Also include teachers with partial records (only check-in, no check-out)
            # They will show 0 days and 0 hours
            for rec in all_records_for_daily:
                teacher_id = rec['teacher_id']
                check_in = rec.get('check_in', '').strip() if rec.get('check_in') else ''
                check_out = rec.get('check_out', '').strip() if rec.get('check_out') else ''
                
                # If teacher has only check-in (no check-out), add them with 0 stats
                if check_in and not check_out:
                    if teacher_id not in teacher_summary:
                        teacher_summary[teacher_id] = {
                            'name': rec['name'],
                            'department': rec['department'],
                            'days': 0,
                            'total_minutes': 0
                        }
            
            # Add summary rows
            for teacher_id, summary in teacher_summary.items():
                # Calculate totals
                total_hours_str = format_minutes_to_hours(summary['total_minutes']) if summary['days'] > 0 else '0 hours 0 minutes'
                avg_minutes = summary['total_minutes'] // summary['days'] if summary['days'] > 0 else 0
                avg_hours_str = format_minutes_to_hours(avg_minutes) if summary['days'] > 0 else '0 hours 0 minutes'
                
                summary_ws.append([
                    summary['name'],
                    summary['department'],
                    summary['days'],
                    total_hours_str,
                    avg_hours_str
                ])
            
            # Sort summary by name
            if summary_ws.max_row > 1:
                # Get all data rows (skip header)
                data_rows = []
                for row in summary_ws.iter_rows(min_row=2, values_only=True):
                    data_rows.append(row)
                
                # Sort by name (first column)
                data_rows.sort(key=lambda x: x[0] or '')
                
                # Clear and re-add sorted rows
                summary_ws.delete_rows(2, summary_ws.max_row)
                for row in data_rows:
                    summary_ws.append(row)
            
            # Set column widths for summary (5 columns now)
            summary_widths = [25, 18, 18, 20, 20]
            for i, width in enumerate(summary_widths, start=1):
                summary_ws.column_dimensions[chr(64 + i)].width = width
        
        # ============================================
        # SHEET 2+: DAILY SHEETS
        # ============================================
        for date_str in sorted_dates:
            # Create sheet for this date (Excel sheet names have 31 char limit)
            sheet_name = date_str[:31] if len(date_str) <= 31 else date_str[:28] + "..."
            daily_ws = wb.create_sheet(sheet_name)
            
            # Daily headers
            daily_headers = ['Name', 'Department', 'Check-in', 'Check-out', 'Working Hours']
            daily_ws.append(daily_headers)
            
            # Style daily header
            for cell in daily_ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Get records for this date (use all_records_for_daily to include partial records)
            date_records = [r for r in all_records_for_daily if r['date'] == date_str]
            
            # Sort by name
            date_records.sort(key=lambda x: x['name'] or '')
            
            # Add rows for this date
            for rec in date_records:
                daily_ws.append([
                    rec['name'],
                    rec['department'],
                    rec['check_in'] or '-',
                    rec['check_out'] or '-',
                    rec['working_hours'] or '-'
                ])
            
            # Set column widths for daily sheet
            daily_widths = [25, 18, 15, 15, 20]
            for i, width in enumerate(daily_widths, start=1):
                daily_ws.column_dimensions[chr(64 + i)].width = width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with date range
        if start_date and end_date:
            filename = f"attendance_{start_date}_to_{end_date}.xlsx"
        elif start_date:
            filename = f"attendance_from_{start_date}.xlsx"
        elif end_date:
            filename = f"attendance_until_{end_date}.xlsx"
        else:
            filename = f"attendance_records_{get_date_string()}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        print(f"Error generating Excel file: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': f'Failed to generate Excel file: {str(e)}',
            'server_time': get_server_time()
        }), 500


@app.route('/api/attendance', methods=['GET'])
def api_attendance():
    """
    Public API endpoint to get all attendance records.
    Returns JSON data that can be consumed by external applications.
    
    Query parameters:
    - date: Optional filter by single date (YYYY-MM-DD) - for backward compatibility
    - start_date: Optional filter start date (YYYY-MM-DD)
    - end_date: Optional filter end date (YYYY-MM-DD)
    - teacher_id: Optional filter by teacher_id
    - format: Optional response format ('detailed' or 'summary', default: 'detailed')
    """
    try:
        teachers = database.get_all_teachers()
        
        # Get query parameters
        filter_date = request.args.get('date')  # Single date (backward compatibility)
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        filter_teacher_id = request.args.get('teacher_id')
        response_format = request.args.get('format', 'detailed')
        
        # Flatten attendance into records
        records = []
        for teacher_id, data in teachers.items():
            # Apply teacher_id filter if provided
            if filter_teacher_id and teacher_id != filter_teacher_id:
                continue
            
            name = data.get('name', '')
            department = data.get('department', '')
            fingerprint_id = data.get('fingerprint_id')
            attendance = data.get('attendance', {}) or {}
            
            if not attendance:
                # Only include if no date filter is applied
                if not filter_date and not start_date and not end_date:
                    records.append({
                        'teacher_id': teacher_id,
                        'name': name,
                        'department': department,
                        'fingerprint_id': fingerprint_id,
                        'date': None,
                        'check_in': None,
                        'check_out': None,
                        'working_hours': None
                    })
            else:
                for date_str, rec in attendance.items():
                    # Apply date filters
                    if filter_date and date_str != filter_date:
                        continue
                    if start_date and date_str < start_date:
                        continue
                    if end_date and date_str > end_date:
                        continue
                    
                    record = {
                        'teacher_id': teacher_id,
                        'name': name,
                        'department': department,
                        'fingerprint_id': fingerprint_id,
                        'date': date_str,
                        'check_in': rec.get('check_in'),
                        'check_out': rec.get('check_out'),
                        'working_hours': rec.get('working_hours')
                    }
                    records.append(record)
        
        # Sort by date descending (most recent first)
        records.sort(key=lambda x: (x['date'] or '', x['name'] or ''), reverse=True)
        
        # Format response based on format parameter
        if response_format == 'summary':
            # Return summary statistics
            total_records = len(records)
            total_teachers = len(set(r['teacher_id'] for r in records if r['teacher_id']))
            dates = sorted(set(r['date'] for r in records if r['date']), reverse=True)
            
            return jsonify({
                'status': 'success',
                'summary': {
                    'total_records': total_records,
                    'total_teachers': total_teachers,
                    'date_range': {
                        'earliest': dates[-1] if dates else None,
                        'latest': dates[0] if dates else None
                    }
                },
                'records': records,
                'server_time': get_server_time()
            }), 200
        else:
            # Return detailed records
            return jsonify({
                'status': 'success',
                'count': len(records),
                'records': records,
                'filters': {
                    'date': filter_date,
                    'start_date': start_date,
                    'end_date': end_date,
                    'teacher_id': filter_teacher_id
                },
                'server_time': get_server_time()
            }), 200
    
    except Exception as e:
        print(f"Error getting attendance API: {e}")
        return jsonify({
            'status': 'error',
            'message': f'Failed to get attendance records: {str(e)}',
            'server_time': get_server_time()
        }), 500


@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint."""
    return jsonify({
        'status': 'healthy',
        'server_time': get_server_time()
    }), 200


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 8000))
    app.run(host="0.0.0.0", port=port, debug=False)
